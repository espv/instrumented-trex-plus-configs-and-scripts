
import errno
import json
import os
import re
from collections import OrderedDict

import numpy as np
import numpy_indexed as npi
import seaborn as sns
from kivy.app import App
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.progressbar import ProgressBar
from kivy.uix.scrollview import ScrollView
from matplotlib import pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl_templates import TemplatedWorkbook
from openpyxl_templates.table_sheet import TableSheet
from openpyxl_templates.table_sheet.columns import CharColumn, IntColumn

trace_fn = None

FIRST_trace_id = '1'

np.set_printoptions(edgeitems=30, linewidth=100000,
                    formatter=dict(float=lambda x: "%.3g" % x))


type_dict = {
    "integer": int,
    "float": float,
    "string": str
}


class TraceSheet(TableSheet):
    line_nr = IntColumn()
    trace_id = CharColumn()
    cpu_id = CharColumn()
    thread_id = CharColumn()
    timestamp = CharColumn()
    time_diff = IntColumn()


class TraceIdSheet(TableSheet):
    line_nr = IntColumn()
    trace_idFrom = CharColumn()
    trace_idTo = CharColumn()
    cpu_id = CharColumn()
    thread_id = CharColumn()
    timestamp = IntColumn()
    time_diff = IntColumn()


class TraceWorkBook(TemplatedWorkbook):
    regular_trace_entries = TraceSheet()
    trace_id_trace_entries = TraceIdSheet()


class TraceEntry(object):
    def __init__(self, line_nr=0, trace_id=0, event_type=0, cpu_id=-1, thread_id=0, timestamp=0, cur_prev_time_diff=0, previous_trace_id=0):
        self.line_nr = line_nr
        self.trace_id = trace_id
        self.event_type = event_type
        self.cpu_id = cpu_id
        self.thread_id = thread_id
        self.timestamp = timestamp
        self.cur_prev_time_diff = cur_prev_time_diff
        self.previous_row = previous_trace_id


class Trace(object):
    def __init__(self, trace, output_fn, trace_ids, reverse_possible_trace_event_transitions, traceAttrs):
        self.rows = []
        self.trace = trace
        self.wb = None
        self.output_fn = output_fn
        self.trace_ids = trace_ids
        self.reverse_possible_trace_event_transitions = reverse_possible_trace_event_transitions
        self.traceAttrs = traceAttrs
        self.numpy_rows = None
        self.max = len(self.rows)+len(self.rows)*2
        self.cnt = 0
        self.update_bar_trigger = None

    def get_previous_event(self, this_row, previous_rows):
        if this_row.event_type == 2:
            for prev in self.reverse_possible_trace_event_transitions.get(this_row.trace_id):
                for i, row in enumerate(reversed(previous_rows)):
                    if row.trace_id == prev:
                        return row
        else:
            restart = True
            while restart:
                restart = False
                for i, row in enumerate(reversed(previous_rows)):
                    index = -(i+1)
                    #if row.thread_id == this_row.thread_id and row.event_type == 1:
                    #    del previous_rows[index]  # We delete this because we have passed this event by now
                    #    restart = True  # Restart loop
                    #    break
                    if row.thread_id == this_row.thread_id:  # and row.cpu_id == this_row.cpu_id:
                        previous_row = row
                        del previous_rows[index]
                        return previous_row

        return TraceEntry(timestamp=this_row.timestamp)

    def collect_data(self):
        previous_times = []
        for line_nr, l in enumerate(self.trace):
            split_l = re.split('[\t\n]', l)
            if len(split_l) < len(self.traceAttrs):
                return -1

            try:
                # Depending on the configuration file, the trace event format might be different
                trace_attr = self.traceAttrs['traceId']
                trace_id = str(type_dict[trace_attr['type']](split_l[int(trace_attr['position'])]))
                cpu_attr = self.traceAttrs['cpuId']
                cpu_id = type_dict[cpu_attr['type']](split_l[int(cpu_attr['position'])])
                thread_attr = self.traceAttrs['threadId']
                thread_id = type_dict[thread_attr['type']](split_l[int(thread_attr['position'])])
                timestamp_attr = self.traceAttrs['timestamp']
                timestamp = type_dict[timestamp_attr['type']](split_l[int(timestamp_attr['position'])])
            except ValueError:  # Occurs if any of the casts fail
                return -1

            event_type = self.trace_ids.get(trace_id, {}).get('type', 0)
            # The last two fields are filled in later
            row = TraceEntry(line_nr+1, trace_id, event_type, thread_id, cpu_id, timestamp, 0, "")
            previous_row = self.get_previous_event(row, previous_times)

            previous_time = previous_row.timestamp
            if trace_id == FIRST_trace_id or line_nr == 0:
                previous_time = row.timestamp
            row.cur_prev_time_diff = timestamp-previous_time
            row.previous_row = previous_row

            numFollowing = self.trace_ids.get(trace_id, {"numFollowing": 1})["numFollowing"]

            self.rows.append(row)

            for _ in range(numFollowing):
                previous_times.append(row)

    @staticmethod
    def adjust_col_width(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 3) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    def regular_as_xlsx(self, pb, popup, bl, btn):
        self.wb = Workbook(write_only=True)

        ws = self.wb.create_sheet("Trace")
        self.wb.active = ws

        trace_file_id = re.split('traces/|[.]trace', self.trace.name)[1]
        fn = 'output/'+trace_file_id+'/'+self.output_fn

        try:
            os.mkdir('output/'+trace_file_id)
        except OSError as exc:  # Python >2.5
            if exc.errno == errno.EEXIST and os.path.isdir('output/'+trace_file_id):
                pass
            else:
                raise

        try:
            self.wb.save(fn)
        except FileNotFoundError:
            pass
        self.wb.close()

        self.wb = load_workbook(fn)

        self.cnt = 0

        def update_bar(_):
            if self.cnt >= len(self.rows):
                try:
                    self.wb.save(fn)
                except FileNotFoundError:
                    pass
                self.wb.close()

                self.wb = load_workbook(fn)  # type: Workbook
                ws = self.wb.active  # type: Worksheet
                font = Font(bold=True, size=14)
                a1 = ws['A1']  # type: Cell
                a1.font = font
                a1.value = "Line number"
                b1 = ws['B1']  # type: Cell
                b1.font = font
                b1.value = "Trace ID"
                c1 = ws['C1']  # type: Cell
                c1.font = font
                c1.value = "Thread ID"
                d1 = ws['D1']  # type: Cell
                d1.font = font
                d1.value = "CPU ID"
                e1 = ws['E1']  # type: Cell
                e1.font = font
                e1.value = "Timestamp"
                f1 = ws['F1']  # type: Cell
                f1.font = font
                f1.value = "Timestamp diff"

                self.adjust_col_width(self.wb.active)
                try:
                    self.wb.save(fn)
                except FileNotFoundError:
                    pass
                self.wb.close()
                popup.open()
                self.cnt = self.max
                bl.add_widget(btn, 3)
                bl.clear_widgets([pb])
                return
            else:
                self.update_bar_trigger()

            pb.value = self.cnt
            te = self.rows[self.cnt]
            self.cnt += 1
            row = [te.line_nr, te.trace_id, te.event_type, te.thread_id,  te.cpu_id, te.timestamp, te.cur_prev_time_diff]
            self.wb.active.append(row)

        self.update_bar_trigger = Clock.create_trigger(update_bar, -1)
        Clock.max_iteration = 100

        self.update_bar_trigger()

    def as_plots(self):
        for row in self.rows:
            if row.trace_id == '1' and  row.previous_row.trace_id == '7':
                print(row)
        self.numpy_rows = np.array([[te.trace_id, te.thread_id, te.cpu_id, te.timestamp, te.cur_prev_time_diff, te.previous_row.trace_id, te.event_type] for te in self.rows])
        if len(self.numpy_rows) == 0:
            return
        print(self.numpy_rows)
        print("\n")
        # First group by => to trace ID
        for group in npi.group_by(self.numpy_rows[:, 0]).split(self.numpy_rows):
            # Then group by => from trace ID
            for g2 in npi.group_by(group[:, 5]).split(group[:, :]):
                self.trace_ids.setdefault(str(group[0][0]), {}).setdefault("traced", []).append({"fromTraceId": g2[0][5], "data": g2})

        trace_file_id = re.split('traces/|[.]trace', self.trace.name)[1]
        try:
            os.mkdir('output/'+trace_file_id)
        except OSError as exc:  # Python >2.5
            if exc.errno == errno.EEXIST and os.path.isdir('output/'+trace_file_id):
                pass
            else:
                raise

        y = []
        xticks = []
        for trace_id, v in self.trace_ids.items():
            for d in v.get("traced", []):
                e = [int(d["data"][i][4]) for i in range(len(d["data"])) if d["data"][i][6] != '2']
                if len(e) > 0:
                    y.append(e)
                    xticks.append(str(d["fromTraceId"])+"-"+trace_id)

        fig, ax = plt.subplots(figsize=(30,5))
        x = np.arange(len(xticks))
        plt.xticks(x, xticks)
        ax.plot(x, np.asarray([np.percentile(fifty, 50) for fifty in y]), label='50th percentile')
        ax.plot(x, np.asarray([np.percentile(fourty, 40) for fourty in y]), label='40th percentile')
        ax.plot(x, np.asarray([np.percentile(thirty, 30) for thirty in y]), label='30th percentile')
        ax.plot(x, np.asarray([np.percentile(twenty, 20) for twenty in y]), label='20th percentile')
        ax.plot(x, np.asarray([np.percentile(seventy, 70) for seventy in y]), label='70th percentile')
        ax.plot(x, np.asarray([np.percentile(eighty, 80) for eighty in y]), label='80th percentile')
        ax.plot(x, np.asarray([np.percentile(sixty, 60) for sixty in y]), label='60th percentile')
        ax.plot(x, np.asarray([np.percentile(ten, 10) for ten in y]), label='10th percentile')
        ax.plot(x, np.asarray([np.percentile(one, 1) for one in y]), label='1th percentile')
        ax.plot(x, np.asarray([np.percentile(ninety, 90) for ninety in y]), label='90th percentile')
        plt.title("Processing delay percentiles")
        plt.xlabel("Processing stage")
        plt.ylabel("Processing delay (nanoseconds)")
        fig.savefig('output/'+trace_file_id+'/percentiles.png')
        plt.show()
        plt.cla()

        flattened_y = np.hstack(np.asarray([np.asarray(e) for e in y]).flatten())
        x = []
        for i, e in enumerate(y):
            for _ in e:
                x.append(i)

        plt.title("Processing delay scatter plot")
        plt.xlabel("Processing stage")
        plt.ylabel("Processing delay (nanoseconds)")
        plt.figure(figsize=(30, 5))
        fig = plt.scatter(x, flattened_y).get_figure()

        plt.xticks(range(len(xticks)), xticks)
        fig.savefig('output/'+trace_file_id+'/scatter.png')
        plt.show()

        for toTraceId, v in self.trace_ids.items():
            for g2 in v.get("traced", []):
                try:
                    if g2["fromTraceId"] == "":
                        continue
                    proc_stage = g2["fromTraceId"] + "-" + toTraceId
                    plt.title("Processing delay histogram for processing stage " + proc_stage)
                    plt.xlabel("Processing delay (µs)")
                    plt.ylabel("Number of events")
                    group = np.array([int(r[4])/1000 for r in g2["data"]])
                    ninetyninth_perc = np.percentile(group, 99)
                    group = np.array([r for r in group if r < ninetyninth_perc])
                    sns_plot = sns.distplot(group, bins=20, kde=False)
                    #sns_plot.xaxis.set_major_locator(ticker.MultipleLocator(10))
                    fig = sns_plot.get_figure()
                    fig.savefig('output/'+trace_file_id+'/processing-stage-'+proc_stage+'.png')

                    plt.show()
                    #plt.title('sequence of events and their processing delays')
                    #plt.ylabel('Processing delay (nanoseconds)')
                    #plt.plot(group, linestyle="None", marker="x")
                    #plt.show()
                except np.linalg.LinAlgError:
                    pass


class TraceAnalysisApp(App):
    icon = 'trace-analysis.png'
    title = "Trace analysis"

    def __init__(self):
        super(TraceAnalysisApp, self).__init__()
        content = Button(text='Success', size_hint_y=None, height=30)
        self.popup = Popup(title='Analysis finished', content=content,
                           auto_dismiss=True, size_hint_y=None, height=30)
        content.bind(on_press=self.popup.dismiss)
        error_content = Button(text='Unknown error encountered when parsing trace. Please try a different trace.', size_hint_y=None, height=30)
        self.error_in_trace_popup = Popup(title='Error', content=error_content)
        error_content.bind(on_press=self.error_in_trace_popup.dismiss)
        self.bl = BoxLayout()
        self.root = ScrollView(size_hint=(1, None), size=(Window.width, Window.height))
        self.root.add_widget(self.bl)
        self.possible_trace_event_transitions = {}
        self.reverse_possible_trace_event_transitions = {}
        self.traceAttrs = {}
        self.trace_id_to_CSEW_events = {}
        self.selected_trace_tb = None
        self.trace_file = None
        self.trace = None
        self.trace_ids = {}
        self.fcl = self.fcl = FileChooserListView(
            path=os.path.realpath("trace-configurations/"), dirselect=True)  # type: FileChooserListView
        self.fcl.bind(selection=self.selected_traceid_to_csem_events_map_file)
        self.bl.add_widget(self.fcl)

    def gen_plots(self, _):
        if self.trace is not None:
            self.trace.as_plots()
        self.popup.open()

    def gen_xlsx(self, btn: Button):
        if self.trace is not None:
            pb = ProgressBar(max=len(self.trace.rows)*2, value=0, size_hint_y=None, height=30)
            self.bl.clear_widgets([btn])
            self.bl.add_widget(pb, 3)
            self.trace.regular_as_xlsx(pb, self.popup, self.bl, btn)

    def parse_trace_file(self):
        self.root.clear_widgets()
        gl = GridLayout(cols=1, size_hint_y=None)
        gl.bind(minimum_height=gl.setter('height'))
        self.root.add_widget(gl)

        gl.add_widget(Label(text='Choose what to do with trace '+self.trace_file.name, size_hint_y=None, height=30))
        gen_plots_btn = Button(text="Analyze data and generate plots", size_hint_y=None, height=30)
        gen_plots_btn.bind(on_press=self.gen_plots)
        gl.add_widget(gen_plots_btn)
        gen_xlsx_btn = Button(text="Analyze data and export to excel", size_hint_y=None, height=30)
        gen_xlsx_btn.bind(on_press=self.gen_xlsx)
        gl.add_widget(gen_xlsx_btn)
        decomp_trace_btn = Button(text="Decompress trace", size_hint_y=None, height=30)
        decomp_trace_btn.bind(on_press=self.decompress_trace)
        gl.add_widget(decomp_trace_btn)
        back_btn = Button(text="Back", size_hint_y=None, height=30)
        back_btn.bind(on_press=self.clear_and_select_trace)
        gl.add_widget(back_btn)
        exit_btn = Button(text="Exit", size_hint_y=None, height=30)
        exit_btn.bind(on_press=lambda _: exit(0))
        gl.add_widget(exit_btn)

        self.trace = Trace(self.trace_file, self.trace_file.name.split(".trace")[0]+".xlsx", self.trace_ids, self.reverse_possible_trace_event_transitions, self.traceAttrs)

        if self.trace.collect_data() == -1:
            self.root.clear_widgets()
            self.error_in_trace_popup.open()
            self.select_trace_to_analyze()

    def selected_trace_file(self, _, selection):
        self.trace_file = open(selection[0], 'r')
        self.parse_trace_file()
        return self.bl

    def eid_to_event(self, e, cycles):
        res = self.trace_id_to_CSEW_events.get(e, {}).get('csemEvents', "")

        return res.replace("[CPU_CYCLES]", str(cycles), -1)

    def decompress_trace(self, _):
        if self.trace_file is not None:
            self.trace_file = open('../../traces/'+self.selected_trace_tb.text, 'r')

            output_file = open("output/processed-"+self.selected_trace_tb.text, "w")
            output_file.write("EOD\n")
            i = 0
            for l in self.trace_file:
                line = l.split("\t")
                if len(line) < 2:
                    break

                eid = int(line[0])
                i += 1
                cycles = int(line[3])
                output_file.write(self.eid_to_event(eid, cycles)+"\n")
            output_file.write("H        \n")

        self.popup.open()

    def select_trace_to_analyze(self):
        self.fcl.path = os.path.realpath("../../traces/")
        self.fcl.unbind(selection=self.selected_traceid_to_csem_events_map_file)
        self.fcl.bind(selection=self.selected_trace_file)
        return self.fcl

    def clear_and_select_trace(self, _):
        return self.select_trace_to_analyze()

    def selected_traceid_to_csem_events_map_file(self, _, selection):
        trace_file = open(selection[0], 'r')
        json_data = trace_file.read()

        data = json.JSONDecoder(object_pairs_hook=OrderedDict).decode(json_data)
        self.trace_ids = data['traceIDs']
        for k, v in self.trace_ids.items():
            for a in v["transitions"]:
                self.reverse_possible_trace_event_transitions.setdefault(a, []).append(k)  # Assume that a given trace Id can only be preceeded by one trace Id
        self.traceAttrs = data["traceAttributes"]

        return self.clear_and_select_trace(None)

    def build(self):
        return self.root


if __name__ == '__main__':
    TraceAnalysisApp().run()
